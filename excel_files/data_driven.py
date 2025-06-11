import openpyxl
import pytest

def get_data_from_excel():
    credentials_list = []
    wb = openpyxl.load_workbook("credentials.xlsx")
    ws = wb.active
    numbers_of_rows = ws.max_row
    numbers_of_columns = ws.max_column

    for row in range(2, numbers_of_rows + 1):
        rows_list = []
        for column in range(1, numbers_of_columns + 1):
            rows_list.append(ws.cell(row, column).value)
        credentials_list.append(rows_list)

    return credentials_list

@pytest.mark.parametrize("username, password", get_data_from_excel())
def test_login(username, password):
    print(f"Using username {username} and {password}")