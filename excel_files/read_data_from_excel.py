import openpyxl

wb = openpyxl.load_workbook("credentials.xlsx")
ws = wb.active # or you can use wb['credentials']
total_rows = ws.max_row
total_cols = ws.max_column

for row in range(1, total_rows + 1):
    for column in range(1, total_cols + 1):
        print(ws.cell(row, column).value, end=" ")
    print()